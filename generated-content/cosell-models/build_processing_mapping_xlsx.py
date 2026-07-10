"""
Build the 7-column Excel:
  Model Name | Model Table Name | Reporting Schema | Reporting Table Name | Processing Stream | Processing Schema | Processing Table Name

Strategy:
  1. Load the shortcuts extracted (21 explicit mappings)
  2. For each of the 231 model tables:
     - Get: Model Name, Model Table Name from lookup
     - Reporting Schema, Reporting Table Name = Source Schema, Source Table from lookup
     - Processing info:
        a. If in shortcuts, use the shortcut mapping
        b. Else (default): Stream=CoSell, Schema=CoSellGold, Table=ModelTableName
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load shortcuts
SHORTCUTS_PATH = r"c:\WorkFAST-main\generated-content\cosell-models\shortcuts_mapping.json"
with open(SHORTCUTS_PATH, "r", encoding="utf-8") as f:
    SHORTCUTS = json.load(f)

# Load lookup
LOOKUP_PATH = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-model-table-lookup.xlsx"
from openpyxl import load_workbook

wb_lookup = load_workbook(LOOKUP_PATH)
ws_lookup = wb_lookup["TableLookup"]

# Build rows
rows = []
for ri in range(2, ws_lookup.max_row + 1):  # Skip header
    model_file = ws_lookup.cell(row=ri, column=1).value or ""
    model_name = ws_lookup.cell(row=ri, column=2).value or ""
    model_table = ws_lookup.cell(row=ri, column=3).value or ""
    source_schema = ws_lookup.cell(row=ri, column=5).value or ""
    source_table = ws_lookup.cell(row=ri, column=6).value or ""
    
    # Reporting layer (from lookup)
    reporting_schema = source_schema
    reporting_table = source_table
    
    # Processing layer (from shortcuts or default)
    shortcut_info = SHORTCUTS.get(model_table, None)
    
    if shortcut_info:
        proc_stream = shortcut_info.get("stream", "CoSell")
        proc_schema = shortcut_info.get("schema", "CoSellGold")
        proc_table = shortcut_info.get("table", model_table)
    else:
        # Default: CoSell stream, CoSellGold schema
        proc_stream = "CoSell"
        proc_schema = "CoSellGold"
        proc_table = model_table
    
    rows.append({
        "model_name": model_name,
        "model_table": model_table,
        "reporting_schema": reporting_schema,
        "reporting_table": reporting_table,
        "proc_stream": proc_stream,
        "proc_schema": proc_schema,
        "proc_table": proc_table,
    })

print(f"Built {len(rows)} rows")

# Create new Excel with 7 columns
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "ProcessingMapping"

# Header
headers = [
    "Model Name",
    "Model Table Name",
    "Reporting Schema",
    "Reporting Table Name",
    "Processing Stream",
    "Processing Schema",
    "Processing Table Name"
]

for col, hdr in enumerate(headers, start=1):
    cell = ws_new.cell(row=1, column=col)
    cell.value = hdr
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Freeze header
ws_new.freeze_panes = "A2"

# Data rows
for row_idx, row_data in enumerate(rows, start=2):
    ws_new.cell(row=row_idx, column=1).value = row_data["model_name"]
    ws_new.cell(row=row_idx, column=2).value = row_data["model_table"]
    ws_new.cell(row=row_idx, column=3).value = row_data["reporting_schema"]
    ws_new.cell(row=row_idx, column=4).value = row_data["reporting_table"]
    ws_new.cell(row=row_idx, column=5).value = row_data["proc_stream"]
    ws_new.cell(row=row_idx, column=6).value = row_data["proc_schema"]
    ws_new.cell(row=row_idx, column=7).value = row_data["proc_table"]

# Set column widths
col_widths = [25, 35, 20, 30, 25, 20, 30]
for col_idx, width in enumerate(col_widths, start=1):
    ws_new.column_dimensions[get_column_letter(col_idx)].width = width

# Add borders and alternating fills
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
light_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")

for row_idx in range(2, len(rows) + 2):
    for col_idx in range(1, 8):
        cell = ws_new.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = light_fill

# Save
output_path = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-models-processing-mapping.xlsx"
wb_new.save(output_path)

print(f"Saved: {output_path}")
print(f"\nSample rows:")
for r in rows[:5]:
    print(f"  {r['model_table']:35} | {r['proc_stream']:25} | {r['proc_schema']:15} | {r['proc_table']}")
