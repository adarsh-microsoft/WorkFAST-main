"""
Build the corrected 7-column Excel with proper Processing Stream/Schema/TableName.

Matching logic:
- Match "Reporting Table Name" (source_table from lookup) with "ShortcutName" from notebook
- If match found: use notebook's stream/schema/table
- If no match: default to CoSell/CoSellGold/same-as-reporting-table
"""

import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load shortcuts
with open(r"c:\WorkFAST-main\generated-content\cosell-models\shortcuts_final.json", "r") as f:
    SHORTCUTS = json.load(f)

# Build lowercase lookup for case-insensitive matching
SHORTCUTS_LOWER = {k.lower().replace(" ", ""): (k, v) for k, v in SHORTCUTS.items()}

print(f"Loaded {len(SHORTCUTS)} shortcuts")

# Load lookup Excel
wb_lookup = load_workbook(r"c:\WorkFAST-main\generated-content\cosell-models\cosell-model-table-lookup.xlsx")
ws_lookup = wb_lookup["TableLookup"]

# Build rows
rows = []
matches = 0
defaults = 0

for ri in range(2, ws_lookup.max_row + 1):
    model_name = ws_lookup.cell(row=ri, column=2).value or ""
    model_table = ws_lookup.cell(row=ri, column=3).value or ""
    source_schema = ws_lookup.cell(row=ri, column=5).value or ""
    source_table = ws_lookup.cell(row=ri, column=6).value or ""
    
    # Reporting layer
    reporting_schema = source_schema
    reporting_table = source_table
    
    # Try to match reporting_table with shortcuts
    shortcut_info = None
    
    # Try exact match first
    if source_table in SHORTCUTS:
        shortcut_info = SHORTCUTS[source_table]
    # Try without spaces
    elif source_table.replace(" ", "") in SHORTCUTS:
        shortcut_info = SHORTCUTS[source_table.replace(" ", "")]
    # Try case-insensitive
    else:
        key = source_table.lower().replace(" ", "")
        if key in SHORTCUTS_LOWER:
            _, shortcut_info = SHORTCUTS_LOWER[key]
    
    # Also try model_table if source_table didn't match
    if not shortcut_info:
        if model_table in SHORTCUTS:
            shortcut_info = SHORTCUTS[model_table]
        elif model_table.replace(" ", "") in SHORTCUTS:
            shortcut_info = SHORTCUTS[model_table.replace(" ", "")]
        else:
            key = model_table.lower().replace(" ", "")
            if key in SHORTCUTS_LOWER:
                _, shortcut_info = SHORTCUTS_LOWER[key]
    
    if shortcut_info:
        proc_stream = shortcut_info["stream"]
        proc_schema = shortcut_info["schema"]
        proc_table = shortcut_info["table"]
        mapping_source = "Explicit (notebook)"
        matches += 1
    else:
        # Default: CoSell/CoSellGold/same table name
        proc_stream = "CoSell"
        proc_schema = "CoSellGold"
        proc_table = source_table if source_table else model_table
        mapping_source = "Default (CoSell/CoSellGold)"
        defaults += 1
    
    rows.append({
        "model_name": model_name,
        "model_table": model_table,
        "reporting_schema": reporting_schema,
        "reporting_table": reporting_table,
        "proc_stream": proc_stream,
        "proc_schema": proc_schema,
        "proc_table": proc_table,
        "mapping_source": mapping_source,
    })

print(f"Built {len(rows)} rows: {matches} matched, {defaults} defaults")

# Create Excel
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "ProcessingMapping"

# Header
headers = [
    "Model Name",
    "Model Table Name",
    "Reporting Schema",
    "Reporting Table Name",
    "Processing Stream",
    "Processing Schema",
    "Processing Table Name",
]

for col, hdr in enumerate(headers, start=1):
    cell = ws_new.cell(row=1, column=col)
    cell.value = hdr
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Freeze header
ws_new.freeze_panes = "A2"

# Colors
explicit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
default_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")   # Light indigo

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Data rows
for row_idx, row_data in enumerate(rows, start=2):
    ws_new.cell(row=row_idx, column=1).value = row_data["model_name"]
    ws_new.cell(row=row_idx, column=2).value = row_data["model_table"]
    ws_new.cell(row=row_idx, column=3).value = row_data["reporting_schema"]
    ws_new.cell(row=row_idx, column=4).value = row_data["reporting_table"]
    ws_new.cell(row=row_idx, column=5).value = row_data["proc_stream"]
    ws_new.cell(row=row_idx, column=6).value = row_data["proc_schema"]
    ws_new.cell(row=row_idx, column=7).value = row_data["proc_table"]
    
    is_explicit = "Explicit" in row_data["mapping_source"]
    fill = explicit_fill if is_explicit else default_fill
    
    for col_idx in range(1, 8):
        cell = ws_new.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.fill = fill

# Column widths
col_widths = [25, 40, 20, 35, 30, 20, 35]
for col_idx, width in enumerate(col_widths, start=1):
    ws_new.column_dimensions[get_column_letter(col_idx)].width = width

# Save
output_path = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-models-processing-mapping-v2.xlsx"
wb_new.save(output_path)

print(f"\nSaved: {output_path}")

# Show some examples
print("\nSample matched rows:")
for r in rows:
    if "Explicit" in r["mapping_source"]:
        print(f"  {r['model_table']:40} -> {r['proc_stream']:25} {r['proc_schema']:15} {r['proc_table']}")
        if sum(1 for x in rows if "Explicit" in x["mapping_source"]) > 10:
            break

print("\nSample default rows:")
count = 0
for r in rows:
    if "Default" in r["mapping_source"]:
        print(f"  {r['model_table']:40} -> {r['proc_stream']:25} {r['proc_schema']:15} {r['proc_table']}")
        count += 1
        if count >= 5:
            break
