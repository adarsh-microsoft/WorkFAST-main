"""
Parse the Reporting_DataTransfer.Notebook to extract:
  ShortcutName (reporting table) -> {Stream, Schema, Table} (processing layer)

Key logic:
  - Path with LatestPublishedSchemaName -> CoSell stream, CoSellGold schema, table from path
  - Path with literal schema (Silver, Gold, ASP, etc.) -> trace lakehouse ID to stream
  - Only in CoSell, replace LatestPublishedSchemaName -> CoSellGold
"""

import re, json

NOTEBOOK_PATH = r"c:\WorkFAST-main\generated-content\cosell-models\notebook-content.py"

# ---- Define streams and their lakehouse/workspace ID assignments ----
STREAMS = {
    "StreamName_SalesReporting": {"name": "SalesPowerBIReporting", "var_lh": "lakehouseId_SalesReporting", "var_ws": "workspaceId_SalesReporting"},
    "StreamName_PPRReporting": {"name": "PPRReporting", "var_lh": "lakehouseId_PPRReporting", "var_ws": "workspaceId_PPRReporting"},
    "StreamName_AzureReporting": {"name": "AzureReporting", "var_lh": "lakehouseId_AzureReporting", "var_ws": "workspaceId_AzureReporting"},
    "StreamName_PMReporting": {"name": "PartnerMastering_Reporting", "var_lh": "lakehouseId_PMReporting", "var_ws": "workspaceId_PMReporting"},
    "StreamName_PartnerProgramsReporting": {"name": "PartnerPrograms_Reporting", "var_lh": "lakehouseId_PartnerProgramsReporting", "var_ws": "workspaceId_PartnerProgramsReporting"},
    "StreamName_SalesSecurity": {"name": "MSSalesUserSecurity", "var_lh": "lakehouseId_SalesSecurity", "var_ws": "workspaceId_SalesSecurity"},
    "StreamName_IAP": {"name": "GPSIAPReporting", "var_lh": "lakehouseId_IAP", "var_ws": "workspaceId_IAP"},
    "StreamName": {"name": "CoSell", "var_lh": "LakehouseId_Stream", "var_ws": "WorkspaceId_Stream"},
    "StreamName_Reporting": {"name": "Cosell_Reporting", "var_lh": "LakehouseId_reporting", "var_ws": "WorkspaceId_reporting"},
}

# Map variable names to stream names (for detecting which stream a lakehouse belongs to)
VAR_TO_STREAM = {}
for var_key, stream_info in STREAMS.items():
    VAR_TO_STREAM[stream_info["var_lh"]] = stream_info["name"]
    VAR_TO_STREAM[stream_info["var_ws"]] = stream_info["name"]

with open(NOTEBOOK_PATH, "r", encoding="utf-8", errors="replace") as f:
    nb_text = f.read()

# ---- Parse shortcut definitions ----
# Look for patterns: "ShortcutName": "TableName" ... "path": "Tables/.../SourceTableName"
# Extract shortcutDetails lists

SHORTCUTS = {}  # shortcut_name -> {stream, schema, processing_table}

def resolve_path_expr(path_str, var_map):
    """Resolve a path expression like 'Tables/'+LatestPublishedSchemaName+'/'+table to {schema, table}."""
    # Handle case variations (e.g., "Tables/"+"LatestPublishedSchemaName")
    if "LatestPublishedSchemaName" in path_str:
        # It's a CoSell path; extract table name
        # Pattern: "Tables/"+LatestPublishedSchemaName+"/"+"TableName" or similar
        m = re.search(r'["\']?([A-Za-z0-9_]+)["\']?\s*$', path_str)
        table_name = m.group(1) if m else ""
        return {"schema": "CoSellGold", "table": table_name, "stream": "CoSell"}
    
    # Pattern: "Tables/Schema/Table" (literal schema in path)
    parts = re.findall(r'["\']([A-Za-z0-9_./]+)["\']', path_str)
    if parts:
        path = parts[0] if isinstance(parts[0], str) else "".join(parts)
        segs = path.split("/")
        if len(segs) >= 3 and segs[0].lower() == "tables":
            schema = segs[1]
            table_name = segs[-1]  # Last segment is the table
            return {"schema": schema, "table": table_name, "stream": None}  # stream will be resolved via workspace/lakehouse
    
    return None

# Extract shortcut entries using a regex that captures JSON-like structures
# Pattern: "ShortcutName": "X", ... "path": "Y"
shortcut_blocks = re.finditer(
    r'"ShortcutName"\s*:\s*["\']?([A-Za-z0-9_]+)["\']?,.*?"path"\s*:\s*["\']?(f?"?[^"\']+["\']?)',
    nb_text,
    re.DOTALL | re.IGNORECASE
)

for match in shortcut_blocks:
    shortcut_name = match.group(1).strip()
    path_expr = match.group(2).strip()
    
    # Try to resolve path
    path_info = resolve_path_expr(path_expr, VAR_TO_STREAM)
    if path_info:
        SHORTCUTS[shortcut_name] = path_info

# Also handle explicit patterns like "Tables/Gold/DimSalesTime"
pattern_explicit = re.compile(r'"ShortcutName"\s*:\s*["\']([A-Za-z0-9_]+)["\'].*?'
                               r'"path"\s*:\s*f?"Tables/([A-Za-z0-9_]+)/([A-Za-z0-9_]+)"',
                               re.DOTALL | re.IGNORECASE)
for match in pattern_explicit.finditer(nb_text):
    shortcut_name = match.group(1)
    schema = match.group(2)
    table = match.group(3)
    # Trace schema to stream
    stream = None
    for s in STREAMS.values():
        if schema.lower() in [s["name"].lower(), "gold", "silver", "asp"]:
            stream = s["name"]
            break
    if not stream:
        stream = "Unknown"
    SHORTCUTS[shortcut_name] = {"schema": schema, "table": table, "stream": stream}

# ---- Now match against the lookup Excel to build the 7-col mapping ----
INV_PATH = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-model-table-lookup.xlsx"
from openpyxl import load_workbook

wb = load_workbook(INV_PATH)
ws = wb["TableLookup"]

# Build mapping: reporting_table_name -> processing info from SHORTCUTS
# We need to iterate through the lookup and match model table to shortcut name

result_rows = []
for ri in range(2, ws.max_row + 1):  # Skip header
    model_file = ws.cell(row=ri, column=1).value
    model_name = ws.cell(row=ri, column=2).value
    model_table = ws.cell(row=ri, column=3).value
    source_schema = ws.cell(row=ri, column=5).value
    source_table = ws.cell(row=ri, column=6).value
    
    # The model_table might be a shortcut name in the reporting layer
    # Try to find it in SHORTCUTS
    shortcut_key = model_table.strip() if model_table else ""
    proc_info = SHORTCUTS.get(shortcut_key, None)
    
    if proc_info:
        proc_stream = proc_info.get("stream", "")
        proc_schema = proc_info.get("schema", "")
        proc_table = proc_info.get("table", "")
    else:
        # Try case-insensitive search
        for skey, sinfo in SHORTCUTS.items():
            if skey.lower() == shortcut_key.lower():
                proc_stream = sinfo.get("stream", "")
                proc_schema = sinfo.get("schema", "")
                proc_table = sinfo.get("table", "")
                break
        else:
            proc_stream = proc_schema = proc_table = ""
    
    result_rows.append({
        "model_name": model_name or "",
        "model_table": model_table or "",
        "reporting_schema": source_schema or "",
        "reporting_table": source_table or "",
        "proc_stream": proc_stream,
        "proc_schema": proc_schema,
        "proc_table": proc_table,
    })

# Save to JSON for next step
out_json = r"c:\WorkFAST-main\generated-content\cosell-models\mapping.json"
with open(out_json, "w", encoding="utf-8") as f:
    json.dump({
        "shortcuts_found": len(SHORTCUTS),
        "rows": result_rows,
    }, f, indent=2)

print(f"Extracted {len(SHORTCUTS)} shortcuts")
print(f"Mapped {len(result_rows)} model tables to processing layer")
print(f"Output: {out_json}")

# Print sample
for r in result_rows[:5]:
    print(f"  {r['model_table']:30} -> {r['proc_stream']:25} {r['proc_schema']:15} {r['proc_table']}")
