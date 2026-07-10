"""Verify the shortcut mappings in the Excel."""

import json
from openpyxl import load_workbook

# Load shortcuts
with open(r'c:\WorkFAST-main\generated-content\cosell-models\shortcuts_mapping.json') as f:
    shortcuts = json.load(f)

print('Shortcuts found:', len(shortcuts))
print()
print('Shortcut mappings:')
for name in sorted(shortcuts.keys())[:15]:
    info = shortcuts[name]
    print(f"  {name:40} -> {info['stream']:30} {info['schema']:15} {info['table']}")

print()
print('Finding shortcuts in Excel...')
wb = load_workbook(r'c:\WorkFAST-main\generated-content\cosell-models\cosell-models-processing-mapping.xlsx')
ws = wb.active

shortcut_rows = []
for r in range(2, ws.max_row + 1):
    model_table = ws.cell(r, 2).value
    proc_stream = ws.cell(r, 5).value
    proc_schema = ws.cell(r, 6).value
    proc_table = ws.cell(r, 7).value
    
    if model_table in shortcuts:
        expected = shortcuts[model_table]
        if (proc_stream != expected['stream'] or 
            proc_schema != expected['schema'] or 
            proc_table != expected['table']):
            print(f'  MISMATCH: {model_table}')
            print(f'    Expected: {expected}')
            print(f'    Got: stream={proc_stream}, schema={proc_schema}, table={proc_table}')
        else:
            shortcut_rows.append((model_table, proc_stream, proc_schema, proc_table))

print()
print(f'Found {len(shortcut_rows)} shortcuts correctly mapped in Excel:')
for name, stream, schema, table in shortcut_rows[:10]:
    print(f"  {name:40} -> {stream:30} {schema:15} {table}")
