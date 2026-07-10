"""
Rebuild the Excel with an additional column marking which tables have explicit shortcut definitions.
This clarifies data quality: explicit vs. inferred mappings.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load shortcuts
SHORTCUTS_PATH = r"c:\WorkFAST-main\generated-content\cosell-models\shortcuts_mapping.json"
with open(SHORTCUTS_PATH, "r", encoding="utf-8") as f:
    SHORTCUTS = json.load(f)

# Load lookup
LOOKUP_PATH = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-model-table-lookup.xlsx"
from openpyxl import load_workbook

wb_lookup = load_workbook(LOOKUP_PATH)
ws_lookup = wb_lookup["TableLookup"]

# Build rows
rows = []
shortcut_count = 0
for ri in range(2, ws_lookup.max_row + 1):  # Skip header
    model_file = ws_lookup.cell(row=ri, column=1).value or ""
    model_name = ws_lookup.cell(row=ri, column=2).value or ""
    model_table = ws_lookup.cell(row=ri, column=3).value or ""
    source_schema = ws_lookup.cell(row=ri, column=5).value or ""
    source_table = ws_lookup.cell(row=ri, column=6).value or ""
    
    # Reporting layer (from lookup)
    reporting_schema = source_schema
    reporting_table = source_table
    
    # Processing layer (from shortcuts or default)
    # Try to match model_table or source_table to a shortcut
    shortcut_match = None
    for sc_name, sc_info in SHORTCUTS.items():
        if model_table == sc_name or source_table == sc_name or source_table.replace(" ", "") == sc_name:
            shortcut_match = sc_info
            shortcut_count += 1
            break
    
    if shortcut_match:
        proc_stream = shortcut_match.get("stream", "CoSell")
        proc_schema = shortcut_match.get("schema", "CoSellGold")
        proc_table = shortcut_match.get("table", model_table)
        mapping_source = "Explicit (notebook)"
    else:
        # Default: CoSell stream, CoSellGold schema
        proc_stream = "CoSell"
        proc_schema = "CoSellGold"
        proc_table = model_table
        mapping_source = "Default (CoSell/CoSellGold)"
    
    rows.append({
        "model_name": model_name,
        "model_table": model_table,
        "reporting_schema": reporting_schema,
        "reporting_table": reporting_table,
        "proc_stream": proc_stream,
        "proc_schema": proc_schema,
        "proc_table": proc_table,
        "mapping_source": mapping_source,
    })

print(f"Built {len(rows)} rows")
print(f"Shortcut matches: {shortcut_count}")

# Create new Excel with 8 columns (added mapping source)
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "ProcessingMapping"

# Header
headers = [
    "Model Name",
    "Model Table Name",
    "Reporting Schema",
    "Reporting Table Name",
    "Processing Stream",
    "Processing Schema",
    "Processing Table Name",
    "Mapping Source"
]

for col, hdr in enumerate(headers, start=1):
    cell = ws_new.cell(row=1, column=col)
    cell.value = hdr
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Freeze header
ws_new.freeze_panes = "A2"

# Data rows
explicit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
default_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")   # Light indigo

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row_idx, row_data in enumerate(rows, start=2):
    ws_new.cell(row=row_idx, column=1).value = row_data["model_name"]
    ws_new.cell(row=row_idx, column=2).value = row_data["model_table"]
    ws_new.cell(row=row_idx, column=3).value = row_data["reporting_schema"]
    ws_new.cell(row=row_idx, column=4).value = row_data["reporting_table"]
    ws_new.cell(row=row_idx, column=5).value = row_data["proc_stream"]
    ws_new.cell(row=row_idx, column=6).value = row_data["proc_schema"]
    ws_new.cell(row=row_idx, column=7).value = row_data["proc_table"]
    ws_new.cell(row=row_idx, column=8).value = row_data["mapping_source"]
    
    # Apply row coloring based on mapping source
    is_explicit = "Explicit" in row_data["mapping_source"]
    fill = explicit_fill if is_explicit else default_fill
    
    for col_idx in range(1, 9):
        cell = ws_new.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.fill = fill

# Set column widths
col_widths = [25, 35, 20, 30, 25, 20, 30, 30]
for col_idx, width in enumerate(col_widths, start=1):
    ws_new.column_dimensions[get_column_letter(col_idx)].width = width

# Save
output_path = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-models-processing-mapping.xlsx"
wb_new.save(output_path)

print(f"Saved: {output_path}")
print()
print("Legend:")
print("  Light Green (Explicit) = mapped from notebook shortcuts")
print("  Light Indigo (Default)  = inferred as CoSell/CoSellGold (no explicit shortcut)")
print()
print("Sample rows:")
for r in rows[:5]:
    print(f"  {r['model_table']:35} | {r['proc_stream']:25} | {r['mapping_source']}")
