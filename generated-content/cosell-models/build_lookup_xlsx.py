"""
Build an AGENT-FRIENDLY Excel lookup from the CoSell BIM inventory.

Goal: given ANY table name as it appears inside a semantic model, an agent can
instantly resolve WHICH physical table (schema + table) to actually check/query.

Scope = the same models currently shown in the HTML report.
"""
import json, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

INV = r"c:\WorkFAST-main\generated-content\cosell-models\inventory.json"
OUT = r"c:\WorkFAST-main\generated-content\cosell-models\cosell-model-table-lookup.xlsx"

# Keep in sync with the HTML report scope
KEEP_MODELS = {
    "CoMarketingModel.bim",
    "CoSellSemanticModel.bim",
    "Partner Planning and Transition Dataset.bim",
    "PartnerSharingModel.bim",
}

with open(INV, "r", encoding="utf-8") as f:
    inv = json.load(f)

models = [m for m in inv["models"] if m["bimFile"] in KEEP_MODELS]

# ---- Flatten to lookup rows ----
rows = []
for m in models:
    for t in m["tables"]:
        schema = t.get("schema") or ""
        src = t.get("sourceTable") or ""
        slt = t.get("sourceLineageTag") or ""
        calc = bool(t.get("isCalculated"))
        rows.append({
            "model_file": m["bimFile"],
            "model_name": m["modelName"],
            "model_table": t.get("table") or "",
            "norm": (t.get("table") or "").strip().lower(),
            "schema": schema,
            "source_table": src,
            "fq_brackets": slt if slt else ("" if calc else (f"[{schema}].[{src}]" if schema and src else "")),
            "fq_plain": (f"{schema}.{src}" if schema and src else ""),
            "mode": t.get("partitionMode") or t.get("sourceType") or "",
            "calc": calc,
        })

# Detect table names that occur in more than one model (need model context to disambiguate)
name_counts = {}
for r in rows:
    name_counts[r["norm"]] = name_counts.get(r["norm"], 0) + 1
for r in rows:
    r["unique_across_models"] = "Yes" if name_counts[r["norm"]] == 1 else "No"

rows.sort(key=lambda r: (r["model_file"].lower(), r["model_table"].lower()))

# ---- Styling helpers ----
BRAND = "0A3D8F"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
TITLE_FONT = Font(color=BRAND, bold=True, size=16, name="Segoe UI")
SUB_FONT = Font(color="555555", size=10, name="Segoe UI")
CALC_FILL = PatternFill("solid", fgColor="FFF4E0")
MONO = Font(name="Consolas", size=10)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D8DEE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols, header_row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = BORDER

def add_table(ws, name, first_row, last_row, last_col):
    ref = f"A{first_row}:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)

wb = Workbook()

# =========================================================
# Sheet 1: ReadMe
# =========================================================
ws = wb.active
ws.title = "ReadMe"
ws.sheet_view.showGridLines = False
ws["A1"] = "CoSell Semantic Models — Agent Table Lookup"
ws["A1"].font = TITLE_FONT
meta = inv.get("meta", {})
info = [
    "",
    "PURPOSE",
    "  Given ANY table name as it appears inside a semantic model, find the PHYSICAL",
    "  table (schema + table) the agent should actually check / query.",
    "",
    "SOURCE",
    f"  Repo: CoSell  |  Project: Global Partner Solutions  |  Org: mcapsdataengineering",
    f"  Branch: master  |  Folder: /Model",
    f"  Derived from each table's table-level sourceLineageTag = [Schema].[Table].",
    f"  Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
    "",
    "SHEETS",
    "  1. TableLookup      Forward lookup. One row per (Model, Model Table).",
    "                      Start here: match 'Model Table Name' (case-insensitive).",
    "  2. BySourceTable    Reverse lookup. One row per (Source Table, consuming Model Table).",
    "                      Use for impact analysis: 'who consumes [Schema].[Table]?'",
    "  3. Models           Per-model summary (table counts, schemas).",
    "",
    "HOW AN AGENT SHOULD USE IT",
    "  1. Normalize the model table name you were given (trim + lower-case).",
    "  2. Look it up in TableLookup against 'Match Key (lower)'.",
    "  3. If 'Unique Across Models' = No, you MUST also know the model to disambiguate",
    "     (the same table name exists in more than one model).",
    "  4. Read 'Source Schema' + 'Source Table'. Query target = 'Fully Qualified ([Schema].[Table])'.",
    "  5. If 'Is Calculated' = Yes, there is NO physical source table — it is computed in",
    "     the model (DAX). Do not look for it in the database.",
    "",
    "COLUMN GLOSSARY (TableLookup)",
    "  Model File              .bim file name in /Model.",
    "  Model Name              Internal model name (can differ from file name).",
    "  Model Table Name        Table name as shown in the semantic model (may contain spaces).",
    "  Match Key (lower)       Lower-cased, trimmed table name for exact programmatic matching.",
    "  Source Schema           Physical schema to check.",
    "  Source Table            Physical table to check.",
    "  Fully Qualified         [Schema].[Table] — verbatim sourceLineageTag.",
    "  Source (plain)          schema.table — convenient for SQL.",
    "  Partition Mode          directLake / import / calculated.",
    "  Is Calculated           Yes = no physical source (computed in model).",
    "  Unique Across Models    No = name appears in >1 in-scope model (needs model context).",
]
r = 2
for line in info:
    ws.cell(row=r, column=1, value=line).font = SUB_FONT if line and not line.isupper() else Font(bold=True, color=BRAND, size=11, name="Segoe UI")
    r += 1
ws.column_dimensions["A"].width = 95

# =========================================================
# Sheet 2: TableLookup
# =========================================================
ws = wb.create_sheet("TableLookup")
ws.sheet_view.showGridLines = False
headers = ["Model File", "Model Name", "Model Table Name", "Match Key (lower)",
           "Source Schema", "Source Table", "Fully Qualified ([Schema].[Table])",
           "Source (plain)", "Partition Mode", "Is Calculated", "Unique Across Models"]
ws.append(headers)
style_header(ws, len(headers))
ws.freeze_panes = "A2"
for r in rows:
    ws.append([
        r["model_file"], r["model_name"], r["model_table"], r["norm"],
        r["schema"] or ("—" if r["calc"] else ""),
        r["source_table"] or ("—" if r["calc"] else ""),
        r["fq_brackets"] or "—",
        r["fq_plain"] or "—",
        r["mode"] or "",
        "Yes" if r["calc"] else "No",
        r["unique_across_models"],
    ])
# mono + wrap + highlight calculated rows
last = ws.max_row
for ri in range(2, last + 1):
    is_calc = ws.cell(row=ri, column=10).value == "Yes"
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.alignment = TOP
        cell.border = BORDER
        if ci in (4, 7, 8):
            cell.font = MONO
        if is_calc:
            cell.fill = CALC_FILL
add_table(ws, "TableLookup", 1, last, len(headers))
widths = [34, 26, 34, 30, 16, 30, 34, 30, 14, 13, 18]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Sheet 3: BySourceTable (reverse lookup)
# =========================================================
ws = wb.create_sheet("BySourceTable")
ws.sheet_view.showGridLines = False
rheaders = ["Source Schema", "Source Table", "Fully Qualified ([Schema].[Table])",
            "Consumed By Model", "As Model Table", "Partition Mode"]
ws.append(rheaders)
style_header(ws, len(rheaders))
ws.freeze_panes = "A2"
rev = [r for r in rows if not r["calc"] and r["schema"] and r["source_table"]]
rev.sort(key=lambda r: (r["schema"].lower(), r["source_table"].lower(), r["model_file"].lower()))
for r in rev:
    ws.append([r["schema"], r["source_table"], r["fq_brackets"],
               r["model_file"], r["model_table"], r["mode"] or ""])
last = ws.max_row
for ri in range(2, last + 1):
    for ci in range(1, len(rheaders) + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.alignment = TOP
        cell.border = BORDER
        if ci == 3:
            cell.font = MONO
add_table(ws, "BySourceTable", 1, last, len(rheaders))
for i, w in enumerate([16, 32, 34, 34, 34, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Sheet 4: Models summary
# =========================================================
ws = wb.create_sheet("Models")
ws.sheet_view.showGridLines = False
mheaders = ["Model File", "Model Name", "Tables", "Source-Mapped", "Calculated", "Schemas"]
ws.append(mheaders)
style_header(ws, len(mheaders))
ws.freeze_panes = "A2"
for m in sorted(models, key=lambda x: -x["tableCount"]):
    mapped = sum(1 for t in m["tables"] if t.get("sourceLineageTag"))
    calc = sum(1 for t in m["tables"] if t.get("isCalculated"))
    schemas = sorted({t.get("schema") for t in m["tables"] if t.get("schema")})
    ws.append([m["bimFile"], m["modelName"], m["tableCount"], mapped, calc, ", ".join(schemas)])
last = ws.max_row
for ri in range(2, last + 1):
    for ci in range(1, len(mheaders) + 1):
        ws.cell(row=ri, column=ci).border = BORDER
        ws.cell(row=ri, column=ci).alignment = TOP
add_table(ws, "ModelsSummary", 1, last, len(mheaders))
for i, w in enumerate([40, 30, 10, 14, 12, 40], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)

print("Workbook written:", OUT)
print("Models:", len(models))
print("TableLookup rows:", len(rows))
print("BySourceTable rows:", len(rev))
amb = sum(1 for r in rows if r["unique_across_models"] == "No")
print("Ambiguous (name in >1 model) rows:", amb)
