"""Generate an Excel workbook of one-line table definitions for the
GPSMartTableDefinitions sheet (CoSell GPS mart).

Output: generated-content/GPSMart_TableDefinitions.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (Table Name, One-line definition) in the exact order supplied.
ROWS = [
    ("Associated_DimPartnerDeal", "Associative dimension linking partner deals to a primary entity (e.g., opportunities) to support many-to-many partner-deal associations."),
    ("Associated_DimPartnerReferral", "Associative dimension linking partner referrals to a primary entity to support many-to-many referral relationships."),
    ("BridgeCoSellPrioritizedSolutionIndustryCountry", "Bridge table resolving the many-to-many relationship between co-sell prioritized solutions and their target industry/country combinations."),
    ("BridgeCoSellPrioritizedSolutionPracticeCountry", "Bridge table linking co-sell prioritized solutions to partner practice/country combinations."),
    ("CustomerSecurity", "Row-level security table defining which users/roles may access specific customer records."),
    ("DimATU", "Dimension of Account Team Units (ATU), Microsoft's account-aligned sales team grouping."),
    ("DimAccountGeographyHierarchyReporting", "Reporting dimension exposing the account geography hierarchy (area/region/subsidiary/country)."),
    ("DimAccountTagsReporting", "Reporting dimension holding account tags/classifications used for filtering and segmentation."),
    ("DimCRMAccounts", "Dimension of CRM (Dynamics/MSX) customer accounts."),
    ("DimCRMPartnerAccount", "Dimension of partner organizations (partner accounts) sourced from CRM."),
    ("DimCRMPartnerAccountReporting", "Reporting-optimized view of CRM partner accounts."),
    ("DimCRMSolution", "Dimension of solutions as defined in CRM."),
    ("DimCRMSolutionArea", "Dimension of CRM solution areas (e.g., Azure, Modern Work, Security, Business Applications)."),
    ("DimCapacityGeography", "Dimension of geography attributes used for partner capacity planning."),
    ("DimChannelPartner", "Dimension of channel partners (resellers/distributors)."),
    ("DimCosellPrioritizedPartners", "Dimension listing partners prioritized for co-sell motions."),
    ("DimCustomPartnerDealDirectionReporting", "Reporting dimension classifying partner-deal direction (inbound/outbound sharing)."),
    ("DimCustomerGeography", "Dimension of customer geography attributes."),
    ("DimCustomerReporting", "Reporting dimension of customers."),
    ("DimCustomerSegment", "Dimension of customer segments (e.g., Enterprise, Corporate, SMB)."),
    ("DimCustomgtmReadiness", "Dimension capturing custom go-to-market (GTM) readiness attributes."),
    ("DimDeliverables", "Dimension of deliverables associated with partner engagements/programs."),
    ("DimDigitallyOffboardedPartner", "Dimension flagging partners that have been digitally offboarded."),
    ("DimEngagementMilestone", "Dimension of engagement milestones in the co-sell/partner lifecycle."),
    ("DimEngagementMilestoneReporting", "Reporting view of engagement milestones."),
    ("DimFiscalMonthReporting", "Fiscal calendar dimension at fiscal-month grain for reporting."),
    ("DimGPSBusinessUnit", "Dimension of GPS (Global Partner Solutions) business units."),
    ("DimGPSCRMAccount", "Dimension of GPS-scoped CRM accounts."),
    ("DimGPSCRMAccountTag", "Dimension of tags applied to GPS CRM accounts."),
    ("DimGPSCRMAccount_Monthly", "Monthly snapshot dimension of GPS CRM accounts."),
    ("DimIOPO", "Dimension of Insertion Order / Purchase Order (IO/PO) records."),
    ("DimIPPartner", "Dimension of IP (intellectual property) / ISV partners."),
    ("DimISVConnectApp", "Dimension of ISV Connect applications/offers."),
    ("DimIndustry", "Dimension of industries/verticals."),
    ("DimInvoice", "Dimension of invoices."),
    ("DimLead", "Dimension of leads."),
    ("DimLifecycleStage", "Dimension of lifecycle stages for opportunities/deals."),
    ("DimMPL", "Dimension of Managed Partner List (MPL) partners."),
    ("DimMSXCustomer", "Dimension of MSX (Microsoft Sales Experience) customers."),
    ("DimMSXPartnerSharingCategorization", "Dimension categorizing MSX partner-sharing types."),
    ("DimMSXPartnerSharingCategorizationReporting", "Reporting view of MSX partner-sharing categorization."),
    ("DimMSXProduct", "Dimension of MSX products."),
    ("DimMSXSolutionArea", "Dimension of MSX solution areas."),
    ("DimMSXWinsWithSolutionAttach", "Dimension of MSX wins that have an attached solution."),
    ("DimMSXWorkLoad", "Dimension of MSX workloads."),
    ("DimMarketPlacePublisher", "Dimension of commercial marketplace publishers."),
    ("DimMarketplaceInvoice", "Dimension of marketplace invoices."),
    ("DimMarketplaceOffer", "Dimension of commercial marketplace offers."),
    ("DimOpportunity", "Dimension of sales opportunities."),
    ("DimOpportunityFeedData", "Dimension capturing opportunity feed-data attributes."),
    ("DimOpportunityMilestoneEstCompletionDate", "Dimension of estimated completion dates for opportunity milestones."),
    ("DimOpportunityReporting", "Reporting view of opportunities."),
    ("DimPINMetric", "Dimension of PIN (Partner Investment/Incentive) metrics."),
    ("DimPartner", "Core dimension of partners."),
    ("DimPartnerDeal", "Dimension of partner deals (co-sell deal registrations)."),
    ("DimPartnerDealDuration", "Dimension capturing partner-deal duration attributes."),
    ("DimPartnerDealProfile", "Dimension of partner-deal profile attributes."),
    ("DimPartnerDealSolutionAttached", "Dimension of solutions attached to partner deals."),
    ("DimPartnerDealSolutionHistory", "Historical dimension of partner-deal solution attachments."),
    ("DimPartnerDealTeam", "Dimension of teams associated with partner deals."),
    ("DimPartnerDeal_int", "Intermediate/staging dimension for partner deals."),
    ("DimPartnerOneReporting", "Reporting dimension for PartnerOne (partner master)."),
    ("DimPartnerReferralReporting", "Reporting dimension of partner referrals."),
    ("DimPlanProfile", "Dimension of plan profiles."),
    ("DimPricingLevelHierarchyReporting", "Reporting dimension of the pricing-level hierarchy."),
    ("DimPrioritizedAndReadyPartnersConsumptionOpportunity", "Dimension of prioritized/ready partners for consumption opportunities."),
    ("DimPrioritizedAndReadyPartnersOppty", "Dimension of prioritized/ready partners for opportunities."),
    ("DimProductHierarchyReporting", "Reporting dimension of the product hierarchy."),
    ("DimProductReporting", "Reporting dimension of products."),
    ("DimProject", "Dimension of projects."),
    ("DimRedCarpetActivity", "Dimension of Red Carpet (partner onboarding/engagement program) activities."),
    ("DimReportingPartnerOneMedian", "Reporting dimension capturing PartnerOne median metrics."),
    ("DimReportingPartnerOneSubReporting", "Reporting dimension for PartnerOne sub-accounts."),
    ("DimRevSumDivision", "Dimension of revenue-summary divisions."),
    ("DimRevSumHierarchyReporting", "Reporting dimension of the revenue-summary hierarchy."),
    ("DimSalesPlayReporting", "Reporting dimension of sales plays."),
    ("DimSalesProgram", "Dimension of sales programs."),
    ("DimSalesProgramReporting", "Reporting view of sales programs."),
    ("DimScoreCardRecognitionTime", "Dimension of scorecard recognition time periods."),
    ("DimSegmentHierarchyReporting", "Reporting dimension of the customer/sales segment hierarchy."),
    ("DimSellInCountry", "Dimension of sell-in countries."),
    ("DimSellerCoSellIncentive", "Dimension of seller co-sell incentives."),
    ("DimServiceCompGrouping", "Dimension of service competency groupings."),
    ("DimSolution", "Dimension of solutions."),
    ("DimSolutionArea", "Dimension of solution areas."),
    ("DimSolutionAreaDetailReporting", "Reporting dimension of solution-area detail."),
    ("DimSolutionAreaOppty", "Dimension of solution areas at opportunity grain."),
    ("DimSolutionAreaReporting", "Reporting view of solution areas."),
    ("DimSolutionEngagementPipeline", "Dimension of the solution engagement pipeline."),
    ("DimTask", "Dimension of tasks."),
    ("DimTopTierMonthlyStatus", "Dimension of top-tier partner monthly status."),
    ("DimWorkloadReporting", "Reporting dimension of workloads."),
    ("FactAHRFeedAudit", "Audit fact tracking processing of the AHR (Account Hierarchy) feed."),
    ("FactAHRFeedTable", "Fact table holding AHR (Account Hierarchy) feed records."),
    ("FactAccount", "Fact table of account-level metrics."),
    ("FactAzureConsumptionCTCG", "Fact of Azure consumption metrics for the CTCG measure set."),
    ("FactAzureConsumptionP1CGCT", "Fact of Azure consumption metrics for the PartnerOne (P1) CGCT measure set."),
    ("FactCRMPartner", "Fact table of CRM partner metrics."),
    ("FactCSPSolutionBilledOpportunity", "Fact of CSP (Cloud Solution Provider) solution billed opportunities."),
    ("FactCSPSolutionConsumptionOpportunity", "Fact of CSP solution consumption opportunities."),
    ("FactCSPSolutionPartnerDeal", "Fact linking CSP solutions to partner deals."),
    ("FactCoSellTargets", "Fact of co-sell targets/quotas."),
    ("FactEngagementMilestonePipeline", "Fact of the engagement-milestone pipeline."),
    ("FactEngagementMilestonePipeline_int", "Intermediate/staging fact for the engagement-milestone pipeline."),
    ("FactIOPO", "Fact of Insertion Order / Purchase Order (IO/PO) transactions."),
    ("FactISVConnect", "Fact of ISV Connect metrics."),
    ("FactInvoice", "Fact of invoice transactions."),
    ("FactLead", "Fact of lead metrics."),
    ("FactMBSCommercialTargets", "Fact of MBS (Business Applications) commercial targets."),
    ("FactMBSCommercialTargetsExcel", "Excel-sourced fact of MBS commercial targets."),
    ("FactMBSTargets", "Fact of MBS targets."),
    ("FactMSXCombined", "Combined MSX fact of opportunity/pipeline metrics."),
    ("FactMSXPartnerSharing", "Fact of MSX partner-sharing activity."),
    ("FactMarketplaceBilledSales", "Fact of commercial marketplace billed sales."),
    ("FactOpportunitiesSalesCycleDuration", "Fact of opportunity sales-cycle duration."),
    ("FactOpportunity", "Fact of opportunity metrics."),
    ("FactOpportunityProduct", "Fact of opportunity metrics at product grain."),
    ("FactPRACRTargets", "Fact of PRACR (Prioritized/Ready Azure Consumption Revenue) targets."),
    ("FactPartnerDeal", "Fact of partner-deal metrics."),
    ("FactPartnerDealDuration", "Fact of partner-deal duration metrics."),
    ("FactPartnerDeal_FY20", "FY20 fiscal-year partition of the partner-deal fact."),
    ("FactPartnerDeal_FY21", "FY21 fiscal-year partition of the partner-deal fact."),
    ("FactPartnerDeal_FY22", "FY22 fiscal-year partition of the partner-deal fact."),
    ("FactPartnerDeal_FY23", "FY23 fiscal-year partition of the partner-deal fact."),
    ("FactPartnerDeal_FY24", "FY24 fiscal-year partition of the partner-deal fact."),
    ("FactPartnerDeal_FY25", "FY25 fiscal-year partition of the partner-deal fact."),
    ("FactPartnerOne", "Fact of PartnerOne metrics."),
    ("FactPartnerReferralReporting", "Reporting fact of partner referrals."),
    ("FactProject", "Fact of project metrics."),
    ("FactRecruitISVTargets", "Fact of ISV recruitment targets."),
    ("FactSolution", "Fact of solution metrics."),
    ("FactSolutionBilledOpportunity", "Fact of solution billed opportunities."),
    ("FactSolutionCapacityGeography", "Fact of solution capacity by geography."),
    ("FactSolutionConsumptionOpportunity", "Fact of solution consumption opportunities."),
    ("FactSolutionPartnerDeal", "Fact linking solutions to partner deals."),
    ("HistoricalRecruitFlags", "Historical table of partner recruitment flags."),
    ("HistoryAccount", "Historical (SCD) table of accounts."),
    ("HistoryAccountHistory", "Change-history table tracking account record changes over time."),
    ("HistoryAccountToPersonMapping", "Historical mapping of accounts to persons."),
    ("HistoryAccountToPersonMappingHistory", "Change-history table tracking account-to-person mapping changes over time."),
    ("HistorySolution", "Historical (SCD) table of solutions."),
    ("HistorySolutionHistory", "Change-history table tracking solution record changes over time."),
    ("InvestmentAsk", "Table of partner investment asks/requests."),
    ("MapAccountTag", "Mapping table linking accounts to tags."),
    ("MapIndustrySolution", "Mapping table linking industries to solutions."),
    ("MapOfferPartnerDeal", "Mapping table linking marketplace offers to partner deals."),
    ("MapOpportunityPartner", "Mapping table linking opportunities to partners."),
    ("MapOpportunityPartnerDeal", "Mapping table linking opportunities to partner deals."),
    ("MapOpportunitySalesProgram", "Mapping table linking opportunities to sales programs."),
    ("MapOpportunitySolution", "Mapping table linking opportunities to solutions."),
    ("MapOpptyPartnerAttachAndSolutionAttach", "Mapping table linking opportunity partner-attach and solution-attach records."),
    ("MapPartnerDeal", "Mapping table for partner-deal relationships."),
    ("MapPartnerDealPartnerAttachAndSolutionAttach", "Mapping table linking partner-deal partner-attach and solution-attach records."),
    ("MapPartnerOneAccountTag", "Mapping table linking PartnerOne accounts to tags."),
    ("MapPartnerOpportunity", "Mapping table linking partners to opportunities."),
    ("MapPrioritizedDealSharingAndWins", "Mapping table relating prioritized deal sharing to wins."),
    ("MapPrioritizedOppty", "Mapping table of prioritized opportunities."),
    ("MapSolutionPartnerDeal", "Mapping table linking solutions to partner deals."),
    ("MapSolutionPartnerOneReportingPartnerOneSubAccount_Tab", "Mapping table linking solutions to PartnerOne reporting and PartnerOne sub-accounts."),
    ("MapSolutionPractice", "Mapping table linking solutions to partner practices."),
    ("MapSolutionPracticeIndustryCountry", "Mapping table linking solution practices to industry/country combinations."),
    ("MapSolutionPracticeIndustryCountry_Capacity", "Capacity-planning variant of the solution-practice-industry-country mapping."),
    ("MapSolutionSellInCountry", "Mapping table linking solutions to sell-in countries."),
    ("MapSolutionSellerCoSellIncentive_Capacity", "Mapping table linking solutions to seller co-sell incentives for capacity planning."),
    ("PartnerDealSolution_Snapshot", "Snapshot table of partner-deal solution attachments."),
    ("PartnerSecurity", "Row-level security table governing partner data access."),
    ("ProjectedBaselineCTCG", "Projected baseline for CTCG Azure consumption."),
    ("ProjectedBaselineP1CGCT", "Projected baseline for PartnerOne (P1) CGCT Azure consumption."),
    ("SPMAllAssignmentDefinition", "SPM (Sales Performance Management) table of all assignment definitions."),
    ("SPM_ManagedTopParentAccounts_int", "Intermediate SPM table of managed top-parent accounts."),
    ("SecAccountProductMapReporting", "Security table mapping account-product combinations for reporting RLS."),
    ("SecAccountTeamReporting", "Security table enforcing account-team row-level security for reporting."),
    ("SecDimSellerReporting", "Security table enforcing seller-dimension row-level security for reporting."),
    ("SecDistinctSubsidiaryReporting", "Security table enforcing distinct-subsidiary row-level security for reporting."),
    ("SecOpportunityTeamReporting", "Security table enforcing opportunity-team row-level security for reporting."),
    ("SecSellerAccountProductMapReporting", "Security table mapping seller-account-product combinations for reporting RLS."),
    ("SecSellerHierarchyReporting", "Security table enforcing seller-hierarchy row-level security for reporting."),
    ("SecUPNAccountTeamReporting", "Security table enforcing UPN-based account-team row-level security for reporting."),
    ("SecUPNDimSellerReporting", "Security table enforcing UPN-based seller row-level security for reporting."),
    ("SecUPNOpportunityTeamReporting", "Security table enforcing UPN-based opportunity-team row-level security for reporting."),
    ("SecUPNUserSubsidiaryReporting", "Security table enforcing UPN-based user-subsidiary row-level security for reporting."),
    ("SecUserSubsidiaryReporting", "Security table enforcing user-subsidiary row-level security for reporting."),
    ("SnapshotDimOpportunityWeekly", "Weekly snapshot of the opportunity dimension."),
    ("SnapshotFactMSXCombined", "Snapshot of the combined MSX fact."),
    ("SnapshotMapSolutionPriorityScenarioIndustryCountry", "Snapshot mapping of solution priority scenario by industry/country."),
    ("hub_vw_stringmap", "Hub view exposing CRM string-map (option-set label) lookups."),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GPSMartTableDefinitions"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#", "Table Name", "One-Line Definition"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, (name, definition) in enumerate(ROWS, start=1):
        ws.append([idx, name, definition])
        r = idx + 1
        for col in range(1, 4):
            cell = ws.cell(row=r, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 3))
        if idx % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F2F6FB")

    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(2)].width = 52
    ws.column_dimensions[get_column_letter(3)].width = 100
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(ROWS) + 1}"

    out_dir = Path(__file__).resolve().parent.parent / "generated-content"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "GPSMart_TableDefinitions.xlsx"
    wb.save(out_path)
    print(f"Wrote {len(ROWS)} table definitions to: {out_path}")


if __name__ == "__main__":
    main()
